import pandas as pd
from tabulate import tabulate

from utilities import Utils
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os

class ExcelReportManager:
    """
    A class for managing and exporting test results to an Excel file.

    This class provides functionality to append rows of test result data to an
    Excel file and ensures proper formatting and presentation of the data. It
    supports conditional formatting, column width adjustment, and text wrapping
    for better readability.

    Attributes:
        logger: Logger instance for recording log messages.
        lock: Threading lock to ensure thread-safe operations.
        utils: Instance of the Utils class for utility operations.
        all_test_results_list: List to store all test results data.
    """
    def __init__(self, logger, lock):
        """
        Initializes the ExcelReportManager.

        Args:
            logger: Logger instance for recording log messages.
            lock: Threading lock to ensure thread-safe operations.
        """
        self.logger = logger
        self.lock = lock
        self.utils = Utils(self.logger)
        self.all_test_results_list = []
        self.result_folder = self.utils.get_test_result_folder()

    def add_row(self, test_result):
        """
        Adds a new test result row to the list and updates the Excel file.

        If the 'output.xlsx' file exists, the method appends the new row to the
        existing data in the file. Otherwise, it creates a new file and adds the
        row.

        Args:
            test_result (list): A list containing test result details, including:
                - Test case ID
                - Test case description
                - Status (e.g., Passed, Failed)
                - Browser used
                - Execution date
        """
        with self.lock:
            if self.utils.check_if_file_exists(os.path.join(self.result_folder, "output.xlsx")):
                df = pd.read_excel(os.path.join(self.result_folder, "output.xlsx"))
                self.all_test_results_list = df.values.tolist()
                self.all_test_results_list.append(test_result)
                self.export_to_excel()
            else:
                self.all_test_results_list.append(test_result)
                self.export_to_excel()

    def export_to_excel(self):
        """
        Exports the list of test results to an Excel file ('output.xlsx').

        This method performs the following:
        - Converts the test results list into a DataFrame.
        - Applies conditional formatting to highlight rows with failed statuses.
        - Adjusts column widths and enables text wrapping for better readability.
        - Saves the formatted data to an Excel file.
        - Prints the formatted DataFrame to the console.

        Raises:
            Any exceptions related to file handling or Pandas operations.
        """
        # print(self.all_test_results_list)

        df = pd.DataFrame(self.all_test_results_list, columns=['tc_id', 'tc_description', 'Status', 'Browser', 'Executed Date'])
        data_condition = ['Failed', 'Fail', 'FAILED']

        highlighted_rows = df['Status'].isin(data_condition).map({
            True: 'background-color: red',
            False: ''
        })

        styler = df.style.apply(lambda _: highlighted_rows)
        # print(type(styler))
        # print(pd.__version__)
        output_file = os.path.join(self.result_folder, "output.xlsx")
        styler.to_excel(output_file, index=False)

        # Auto-adjust column widths and enable wrapping
        workbook = load_workbook(output_file)
        sheet = workbook.active
        max_width = 80  # Set maximum column width

        for column in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(column[0].column)  # Get column letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = min(max(max_length, len(str(cell.value))), max_width)
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Enable text wrapping
                except:
                    pass
            sheet.column_dimensions[col_letter].width = max_length + 2  # Add padding

        workbook.save(output_file)

        print(tabulate(df, headers='keys', tablefmt='psql', showindex=False))

    def is_not_used(self):
        """
        Placeholder method. Currently not used.
        """
        pass
